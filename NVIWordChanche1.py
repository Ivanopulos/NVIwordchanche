import pandas as pd
import os
import shutil
import zipfile
import re
import openpyxl
import win32com.client
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import comtypes.client
import time

reestr = "C:\\Users\\IMatveev\\PycharmProjects\\NVIwordchanche\\combined_data_with_votes.xlsx"
def move_files(source_dir, next_number):
    target_dir = os.path.join(source_dir, str(next_number))
    if not os.path.exists(target_dir):
        os.makedirs(target_dir)

    for filename in os.listdir(source_dir):
        if f"_{str(next_number)}" in filename:
            source_file = os.path.join(source_dir, filename)
            target_file = os.path.join(target_dir, filename)
            shutil.move(source_file, target_file)
def docx_to_pdf(docx_filename, pdf_filename):
    word = win32com.client.Dispatch('Word.Application')
    doc = word.Documents.Open(docx_filename)
    doc.SaveAs(pdf_filename, FileFormat=17)
    doc.Close()
    word.Quit()
def replace_words_in_excel(excel_file_path, replacements):  # Функция для замены слов в Excel документе
    # Загрузка документа Excel
    workbook = openpyxl.load_workbook(excel_file_path, keep_vba=True)

    # Итерация по всем листам в документе
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Замена слов в ячейке
                    for old_word, new_word in replacements.items():
                        if old_word in cell.value:
                            cell.value = cell.value.replace(old_word, str(new_word))
        # вставка фркйма чтоб макрос понял
        start_row = sheet['AD1'].row
        start_column = sheet['AD1'].column
        for r_idx, row in enumerate(dataframe_to_rows(df1, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, start_column):
                sheet.cell(row=r_idx, column=c_idx, value=value)
    # Сохранение изменений в новый файл
    workbook.save(excel_file_path.replace('.xlsm', '_modified.xlsm'))
def format_as_date(value):
    try:
        parsed = pd.to_datetime(value)
        return parsed.strftime('%d.%m.%Y')
    except:
        return value

def format_as_time(value):
    try:
        parsed = pd.to_datetime(value)
        return parsed.strftime('%H:%M:%S')
    except:
        return value

# Читаем файлы
meeting = pd.read_excel("meeting.xlsx", dtype=str)

# присваиваем номер
meetingbase_df = pd.read_excel('meetingbase.xlsx', dtype=str)
for column in meeting.columns:
    if column not in meetingbase_df.columns:
        meetingbase_df[column] = None
def rows_equal(series1, series2):
    # Сравниваем две Series с учетом NaN значений
    return (series1.fillna(1) == series2.fillna(1)).all() or (series1.isna() & series2.isna()).all()
row_exists = meetingbase_df.apply(lambda row: rows_equal(row[meeting.columns], meeting.iloc[0]), axis=1)
#row_exists = (meetingbase_df[meeting.columns] == meeting.iloc[0]).all(axis=1)

if not row_exists.any():# Проверяем, существует ли уже такая строка в meetingbase_df
    meetingbase_df['номер'] = pd.to_numeric(meetingbase_df['номер'], errors='coerce')
    next_number = meetingbase_df['номер'].max() + 1
    new_row = meeting.iloc[0].to_dict()  # Добавление новой строки
    new_row['номер'] = next_number
    meetingbase_df = pd.concat([meetingbase_df, pd.DataFrame([new_row])], ignore_index=True)
    meetingbase_df.to_excel('meetingbase.xlsx', index=False)
else:
    next_number = meetingbase_df.loc[row_exists, 'номер'].values[0]
# Форматируем столбцы с датами
for column in meeting.columns:
    if "Дата" in column:
        meeting[column] = meeting[column].apply(format_as_date)
    elif "Время" in column:
        meeting[column] = meeting[column].apply(format_as_time)
print(meeting)
combined_data = pd.read_excel(reestr, dtype=str)
for column in combined_data.columns:
    if "Дата" in column:
        combined_data[column] = combined_data[column].apply(format_as_date)

# Объединение данных на основе логина
df = pd.merge(meeting, combined_data, left_on="User Login", right_on="Login", how="inner")
df1 = pd.DataFrame({
    'metka': '{' + df.columns + '}',
    'chenge': df.iloc[0].values
})

agenda_rows = df1[df1['metka'].str.contains('Пункт повестки')]

# Объединяем данные "Пункт повестки" в одну строку
formatted_agenda = ' '.join([    # '{Пункт повестки}
    f"{i + 1}. {row.chenge} </w:t><w:br/><w:t>"
    for i, row in enumerate(agenda_rows.itertuples())
])
formatted_agenda1 = ' '.join([     # '{Пункт повестки с запятыми}
    f'{i + 1}. {row.chenge} ", "'
    for i, row in enumerate(agenda_rows.itertuples())
])
formatted_agenda1 = '"' + formatted_agenda1
formatted_agenda1 = formatted_agenda1[:-3]
df1 = df1.fillna("")

# Удаляем старые строки "Пункт повестки" и добавляем новую объединенную строку
df1 = df1[~df1['metka'].str.contains('Пункт повестки')]
new_row = pd.DataFrame([{'metka': '{Пункт повестки}', 'chenge': formatted_agenda}])
df1 = pd.concat([df1, new_row], ignore_index=True)
new_row = pd.DataFrame([{'metka': '{Пункт повестки с запятыми}', 'chenge': formatted_agenda1}])
df1 = pd.concat([df1, new_row], ignore_index=True)
df1.loc[len(df1)] = {'metka': "{номер голосования}", 'chenge': next_number}

df1.to_excel("asd.xlsx")

pathwords = ("Уведомление_шаблон_для_заполнения.docx", "vote_str.php")  #  "Бюллетень_шаблон.docx"
pathzip = "B.zip"
df = df1  # для совместимости с копией замены по меткам

for pathword in pathwords:
    mem = 0
    isch = 0
    usch = 0
    found = ""

    try:
        os.remove("B.zip")
    except:
        asd = 1
    if pathword[-4:] == "docx":
        shutil.copy(pathword, pathzip)
        fantasy_zip = zipfile.ZipFile(pathzip)  # extract zip (+need rename docx to zip +need raname vise versa
        fantasy_zip.extractall("/B")
        fantasy_zip.close()
        with open("/B/word/document.xml", 'r', encoding='utf-8') as f:
            content = f.read()
        # Применяем замену с использованием регулярного выражения
        content = re.sub(r"(\{[^}{]*?)</w:t>[^}{]*?<w:t>([^}{]*?})", r"\1\2", content) #чтоб каждый раз не удалять лишнее
        # Записываем обновленное содержимое обратно в файл
        with open("/B/word/document.xml", 'w', encoding='utf-8') as f:
            f.write(content)
        doc = "/B/word/document.xml"
    else:
        doc = pathword

    with open(doc, 'r', encoding='utf-8') as f:  # save before chenge
        get_all = f.readlines()
    print("xml opened")

    if not pathword[-4:] == "docx":# меняем имя потому что в доке рождается копия а в других иначе в оригинале будет править
        doc = pathword.split("_")[0] + "_" + str(next_number) + "_" + str(df1.loc[df1['metka'] == '{User Login}', 'chenge'].iloc[0]) + "." + pathword.split('.')[-1]

    with open(doc, 'w', encoding='utf-8') as f:  # look for { and chenge it
        for i in get_all:         # STARTS THE NUMBERING FROM 1 (by default it begins with 0)
            mem = 0
            found = ""
            usch = len(i)-1
            for u in i:
                try:
                    if get_all[isch][usch] == "}":
                        mem = 1
                except:
                    print(isch, usch, u, i, get_all)
                    print(get_all[isch][usch])
                if mem == 1:
                    found = get_all[isch][usch] + found
                if get_all[isch][usch] == "{" and mem == 1:
                    mem = 0
                    #found = re.sub(r"(\{[^}{]*?)</w:t>[^}{]*?<w:t>([^}{]*?})", r"\1\2", found)
                    #found = re.sub(r"</w:t>[^}{]*?(?=})", "", found)
                    print(found)
                    print(df[df["metka"] == found]["chenge"].values[0])
                    tx = df[df["metka"] == found]["chenge"].values[0]
                    try:
                        float(tx)
                        tx = str(tx).replace(".", ",")
                    except:
                        tx = str(tx)
                        asd = 0
                    get_all[isch] = get_all[isch][:usch] + tx + get_all[isch][usch + len(found):]
                    found = ""
                usch = usch - 1
            isch = isch + 1
        f.writelines(get_all)
    print("XML/txt chanched")

    if pathword[-4:] == "docx":
        name = pathword.split("_")[0] + "_" + str(next_number) + "_" + str(df1.loc[df1['metka'] == '{User Login}', 'chenge'].iloc[0])#
        try:
            os.remove("B.zip")
        except:
            asd = 1
        fantasy_zip = zipfile.ZipFile("B.zip", 'w')
        for folder, subfolders, files in os.walk("/B"):
            for file in files:
                fantasy_zip.write(os.path.join(folder, file), os.path.relpath(os.path.join(folder, file), "/B"))
        fantasy_zip.close()  # transform it to zip
        print("zip saved")
        try:
            os.remove(name + ".docx")
            print(name, "removed")
        except:
            asd = 1
        os.rename("B.zip", name + ".docx")
        shutil.rmtree("/B/")
        current_directory = os.getcwd()
        name = os.path.join(current_directory, name)
        docx_to_pdf(name + ".docx", name + ".pdf")


    print("FINISH")



replacements = dict(zip(df1['metka'], df1['chenge']))  # Создание словаря для замены
replace_words_in_excel('psk.xlsm', replacements)  # Замена слов в файле 'psk.xlsm'
# макрос
current_directory = os.getcwd()
full_path = os.path.join(current_directory, 'psk_modified.xlsm')
full_path1 = os.path.join(current_directory, str(next_number), 'psk_modified.xlsm')
excel = win32com.client.Dispatch("Excel.Application")
workbook = excel.Workbooks.Open(full_path)
excel.DisplayAlerts = False
excel.Application.Run("mmmm")
workbook.Close(False)
excel.Application.Quit()

move_files(current_directory, next_number)